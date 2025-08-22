import eel
from openpyxl import Workbook 
from database_handler.db_api import get_session_obj
from database_handler.db_api import read_fans_from_db
from database_handler.db_api import insert_fans_into_db
from database_handler.db_api import update_fan_stock_into_db_add
from database_handler.db_api import update_fan_stock_into_db_reduce
from database_handler.db_api import delete_fan_stock_into_db
from database_handler.db_api import read_fanStock_from_db

session = None


def main():
    session = get_session_obj()
    # end of business logix

    eel.init("web")
    try:

        @eel.expose  # expose this function to javascript
        def say_hello(arg):
            print("ServerSide:", arg)

        @eel.expose  # expose this function to javascript
        def get_all_fan_data():
            return read_fans_from_db(session)
        
        @eel.expose  # expose this function to javascript
        def gen_excel_and_store_in_dir():
            #open CMD
            #pip install openpyxl
            #from openpyxl import Workbook
            wb = Workbook()
            # ws = wb.active
            # ws['A1'] = "ID"
            # ws['B1'] = "Name"
            # ws['A2'] = 1
            # ws['B2'] = "Fan1"
            # wb.save('fileFromPy.xlsx')
            data = read_fans_from_db(session)
            # data = [{
            #         "fan_id": obj.id,
            #         "fan_name": obj.name, 
            #         "fan_rate": obj.rate, 
            #     }]
            wb = Workbook()
            workbookSheet1 = wb.active
            header_list=["FanId", "Name", "Price"]
            workbookSheet1.append(header_list)
            for d in data:
              #   {
              #         "fan_id": obj.id,
              #         "fan_name": obj.name, 
              #         "fan_rate": obj.rate, 
              #     }
              print(d)
              workbookSheet1.append([
                  d["fan_id"],d["fan_name"],d["fan_rate"]
              ])   
            wb.save('fanData.xlsx')   
            return "Excel Generated"
        
        @eel.expose
        def get_all_fanStock_data():
            return read_fanStock_from_db(session)

        @eel.expose
        def inserting_data(fan_dict):
            return insert_fans_into_db(session, fan_dict)

        @eel.expose
        def update_fan_stock_add(data):
            return update_fan_stock_into_db_add(session, data)

        @eel.expose
        def update_fan_stock_reduce(data):
            return update_fan_stock_into_db_reduce(session, data)

        @eel.expose
        def delete_fan_stock(data):
            return delete_fan_stock_into_db(session, data)

        screen_w = 960
        screen_h = 720

        eel.start("home.html", size=(screen_w, screen_h), port=3030)  # start
        print("closed Succcessfully")
    except Exception as e:
        print("error:", e)


if __name__ == "__main__":
    main()
