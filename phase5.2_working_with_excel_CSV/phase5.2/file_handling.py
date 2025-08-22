# open(filepath/name"testFile.txt","w" <= mode)
#filepath: complete file path with Drive+Folder+fileName.extantion
# filename
file_obj = open("testFile.txt","w+")
file_obj.close()

with open("testFile2.txt","w+") as f:
  f.write("2nd file content")


# gen Excel

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "ID"
ws['B1'] = "Name"
ws['A2'] = 1
ws['B2'] = "Fan1"
wb.save('fileFromPy.xlsx')